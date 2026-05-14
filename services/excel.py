from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from models.invoice import Invoice

_HEADERS = ["Name", "Quantity", "Unit", "Price per unit", "Total"]


def generate_poster_excel(invoice: Invoice, filename: str = "supply.xlsx") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Supply"

    header_font = Font(bold=True)
    for col_idx, header in enumerate(_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, item in enumerate(invoice.items, start=2):
        ws.cell(row=row_idx, column=1, value=item.name)
        ws.cell(row=row_idx, column=2, value=item.quantity)
        ws.cell(row=row_idx, column=3, value=item.unit)
        ws.cell(row=row_idx, column=4, value=item.price_per_unit)
        ws.cell(row=row_idx, column=5, value=item.total)

    for col_idx in range(1, len(_HEADERS) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, ws.max_row + 1)
        )
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
